"""
Admin Statistics Controller
Endpoints thống kê cho Admin Dashboard
"""

from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, Query

from app.core.deps import AuthorizationService
from app.schemas.admin.statistics import (
    ActivityStatsResponse,
    CourseStatsResponse,
    FinanceStatsResponse,
    InstructorStatsResponse,
    OverviewResponse,
    RevenueByCategoryResponse,
    RevenueStatsResponse,
    TopCoursesResponse,
    TopInstructorsResponse,
    UserStatsResponse,
)
from app.services.admin.statistics import StatisticsService

router = APIRouter(prefix="/admin/statistics", tags=["ADMIN STATISTICS"])


# ================================================================
# 1. OVERVIEW
# ================================================================
@router.get("/overview", response_model=OverviewResponse)
async def get_overview(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
):
    """
    Lấy thống kê tổng quan cho Dashboard Admin.

    Bao gồm:
    - Tổng users, courses, instructors
    - Doanh thu tổng và hôm nay
    - Số yêu cầu rút tiền/hoàn tiền đang chờ
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_overview_async()


# ================================================================
# 2. REVENUE STATS
# ================================================================
@router.get("/revenue", response_model=RevenueStatsResponse)
async def get_revenue_stats(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    period: str = Query("month", description="Khoảng thời gian: day|week|month|year"),
    from_date: Optional[date] = Query(None, description="Ngày bắt đầu (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="Ngày kết thúc (YYYY-MM-DD)"),
):
    """
    Thống kê doanh thu theo thời gian.

    - **period**: Nhóm theo ngày/tuần/tháng/năm
    - **from_date/to_date**: Lọc theo khoảng thời gian cụ thể
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_revenue_stats_async(period, from_date, to_date)


# ================================================================
# 3. REVENUE BY CATEGORY
# ================================================================
@router.get("/revenue/by-category", response_model=RevenueByCategoryResponse)
async def get_revenue_by_category(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    from_date: Optional[date] = Query(None, description="Ngày bắt đầu"),
    to_date: Optional[date] = Query(None, description="Ngày kết thúc"),
):
    """
    Thống kê doanh thu theo danh mục khóa học.

    Trả về danh sách categories kèm doanh thu và phần trăm.
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_revenue_by_category_async(from_date, to_date)


# ================================================================
# 4. USER STATS
# ================================================================
@router.get("/users", response_model=UserStatsResponse)
async def get_user_stats(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    period: str = Query("month", description="Khoảng thời gian growth: day|week|month"),
):
    """
    Thống kê người dùng.

    Bao gồm:
    - Tổng/verified/banned
    - Phân bổ theo role
    - Tăng trưởng user mới
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_user_stats_async(period)


# ================================================================
# 5. COURSE STATS
# ================================================================
@router.get("/courses", response_model=CourseStatsResponse)
async def get_course_stats(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
):
    """
    Thống kê khóa học.

    Bao gồm:
    - Tổng số khóa học
    - Phân bổ theo status (published/draft/archived)
    - Phân bổ theo level
    - Rating trung bình
    - Tổng enrollments
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_course_stats_async()


# ================================================================
# 6. TOP COURSES
# ================================================================
@router.get("/courses/top", response_model=TopCoursesResponse)
async def get_top_courses(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    sort_by: str = Query(
        "revenue", description="Sắp xếp theo: revenue|views|enrollments"
    ),
    limit: int = Query(10, ge=1, le=50, description="Số lượng kết quả"),
):
    """
    Top khóa học.

    - **sort_by**: Tiêu chí sắp xếp (doanh thu/views/enrollments)
    - **limit**: Số lượng trả về (1-50)
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_top_courses_async(sort_by, limit)


# ================================================================
# 7. INSTRUCTOR STATS
# ================================================================
@router.get("/instructors", response_model=InstructorStatsResponse)
async def get_instructor_stats(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
):
    """
    Thống kê giảng viên.

    Bao gồm:
    - Tổng số giảng viên
    - Tổng earnings/pending/paid
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_instructor_stats_async()


# ================================================================
# 8. TOP INSTRUCTORS
# ================================================================
@router.get("/instructors/top", response_model=TopInstructorsResponse)
async def get_top_instructors(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    sort_by: str = Query(
        "revenue", description="Sắp xếp theo: revenue|students|courses"
    ),
    limit: int = Query(10, ge=1, le=50, description="Số lượng kết quả"),
):
    """
    Top giảng viên.

    - **sort_by**: Tiêu chí sắp xếp (doanh thu/số học viên/số khóa học)
    - **limit**: Số lượng trả về (1-50)
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_top_instructors_async(sort_by, limit)


# ================================================================
# 9. FINANCE STATS
# ================================================================
@router.get("/finance", response_model=FinanceStatsResponse)
async def get_finance_stats(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
):
    """
    Thống kê tài chính.

    Bao gồm:
    - Số dư platform wallet
    - Tổng deposits/withdrawals
    - Pending withdrawals
    - Refund stats
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_finance_stats_async()


# ================================================================
# 10. ACTIVITY STATS
# ================================================================
@router.get("/activity", response_model=ActivityStatsResponse)
async def get_activity_stats(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    period: str = Query("month", description="Khoảng thời gian: day|week|month"),
):
    """
    Thống kê hoạt động.

    Bao gồm:
    - Views
    - Comments
    - Notes
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_activity_stats_async(period)


# ================================================================
# 11. REVENUE COMPARE (Chi tiết)
# ================================================================
@router.get("/revenue/compare")
async def get_revenue_compare(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    current_from: date = Query(..., description="Kỳ hiện tại - Ngày bắt đầu"),
    current_to: date = Query(..., description="Kỳ hiện tại - Ngày kết thúc"),
    previous_from: date = Query(..., description="Kỳ so sánh - Ngày bắt đầu"),
    previous_to: date = Query(..., description="Kỳ so sánh - Ngày kết thúc"),
):
    """
    So sánh doanh thu giữa 2 kỳ.

    Ví dụ: So sánh tháng này với tháng trước.
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_revenue_compare_async(
        current_from, current_to, previous_from, previous_to
    )


# ================================================================
# 12. REVENUE TRANSACTIONS (Chi tiết)
# ================================================================
@router.get("/revenue/transactions")
async def get_revenue_transactions(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    from_date: Optional[date] = Query(None, description="Ngày bắt đầu"),
    to_date: Optional[date] = Query(None, description="Ngày kết thúc"),
    status: Optional[str] = Query(
        None, description="Trạng thái: completed|pending|refunded"
    ),
    page: int = Query(1, ge=1, description="Trang"),
    size: int = Query(20, ge=1, le=100, description="Số lượng mỗi trang"),
):
    """
    Danh sách giao dịch chi tiết có phân trang.

    Lọc theo ngày và trạng thái.
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_revenue_transactions_async(
        from_date, to_date, status, page, size
    )


# ================================================================
# 13. REVENUE BY INSTRUCTOR (Chi tiết)
# ================================================================
@router.get("/revenue/by-instructor")
async def get_revenue_by_instructor(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    from_date: Optional[date] = Query(None, description="Ngày bắt đầu"),
    to_date: Optional[date] = Query(None, description="Ngày kết thúc"),
    limit: int = Query(20, ge=1, le=100, description="Số lượng kết quả"),
):
    """
    Doanh thu chi tiết theo từng giảng viên.

    Bao gồm platform fee, net earning, số giao dịch.
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_revenue_by_instructor_async(from_date, to_date, limit)


# ================================================================
# 14. REVENUE BY COURSE (Chi tiết)
# ================================================================
@router.get("/revenue/by-course")
async def get_revenue_by_course(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    from_date: Optional[date] = Query(None, description="Ngày bắt đầu"),
    to_date: Optional[date] = Query(None, description="Ngày kết thúc"),
    limit: int = Query(20, ge=1, le=100, description="Số lượng kết quả"),
):
    """
    Doanh thu chi tiết theo từng khóa học.

    Bao gồm số lần bán, giá trung bình, số refund.
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_revenue_by_course_async(from_date, to_date, limit)


# ================================================================
# 15. REVENUE TRENDS (Chi tiết)
# ================================================================
@router.get("/revenue/trends")
async def get_revenue_trends(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    period: str = Query("month", description="Nhóm theo: month|week"),
    months: int = Query(12, ge=1, le=36, description="Số tháng lấy data"),
):
    """
    Xu hướng doanh thu theo thời gian.

    Bao gồm growth rate, tháng tốt nhất, trung bình.
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_revenue_trends_async(period, months)


# ================================================================
# 16. REVENUE EXPORT (Xuất báo cáo CSV)
# ================================================================
@router.get("/revenue/export")
async def get_revenue_export(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    group_by: str = Query("day", description="Nhóm theo: day|month|year"),
    from_date: Optional[date] = Query(None, description="Ngày bắt đầu"),
    to_date: Optional[date] = Query(None, description="Ngày kết thúc"),
):
    """
    Xuất báo cáo doanh thu dạng CSV.

    - **group_by**: Nhóm theo ngày/tháng/năm
    - Download file CSV trực tiếp
    """
    import csv
    import io

    await authorization.require_role(["ADMIN"])

    data = await service.get_revenue_export_async(group_by, from_date, to_date)

    # Tên file
    timestamp = date.today().strftime("%Y%m%d")
    filename = f"revenue_{group_by}_{timestamp}.csv"

    # StreamingResponse với headers thủ công
    # Dùng utf-8-sig (BOM) cho Excel

    stream = io.BytesIO()
    # Write BOM
    stream.write(b"\xef\xbb\xbf")

    # Write CSV content
    text_stream = io.StringIO()
    if data:
        # Mapping Header tiếng Việt
        header_map = {
            "date": "Ngày",
            "month": "Tháng",
            "year": "Năm",
            "platform_income": "Doanh thu nền tảng (VND)",
            "instructor_payout": "Trả giảng viên (VND)",
            "total_transaction": "Tổng giá trị (VND)",
            "transaction_count": "Số lượng giao dịch",
        }

        # Get headers based on group_by
        original_keys = list(data[0].keys())

        # Tạo mapping key -> vn_header
        # DictWriter cần fieldnames là keys của dict row (tức là original_keys)
        # Nhưng ta muốn header row hiển thị tiếng Việt.

        writer = csv.DictWriter(text_stream, fieldnames=original_keys)
        # Write header custom (Tiếng Việt)
        writer.writerow(
            dict(zip(original_keys, [header_map.get(k, k) for k in original_keys]))
        )
        # Write data
        writer.writerows(data)

    stream.write(text_stream.getvalue().encode("utf-8"))
    stream.seek(0)

    from fastapi.responses import StreamingResponse

    return StreamingResponse(
        stream,
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ================================================================
# 17. INSTRUCTOR GROWTH
# ================================================================
@router.get("/instructors/growth")
async def get_instructor_growth(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    period: str = Query("month", description="Khoảng thời gian: day|month"),
):
    """
    Biểu đồ tăng trưởng giảng viên.

    - **period**: day (30 ngày gần nhất) hoặc month (12 tháng gần nhất)
    """

    await authorization.require_role(["ADMIN"])
    return await service.get_instructor_growth_async(period)


# ================================================================
# 18. INSTRUCTOR BY CATEGORY
# ================================================================
@router.get("/instructors/by-category")
async def get_instructor_by_category(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    limit_instructors: int = Query(
        5, ge=1, le=20, description="Số top giảng viên mỗi danh mục"
    ),
):
    """
    Thống kê giảng viên theo danh mục khóa học.

    Mỗi danh mục hiển thị:
    - Số giảng viên
    - Số khóa học
    - Tổng doanh thu
    - Top giảng viên theo doanh thu
    """
    await authorization.require_role(["ADMIN"])
    return await service.get_instructor_by_category_async(limit_instructors)


# ================================================================
# 19. INSTRUCTOR EXPORT (Xuất CSV)
# ================================================================
@router.get("/instructors/export")
async def get_instructor_export(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    sort_by: str = Query(
        "revenue", description="Sắp xếp theo: revenue|students|courses"
    ),
    from_date: Optional[date] = Query(None, description="Ngày bắt đầu (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="Ngày kết thúc (YYYY-MM-DD)"),
):
    """
    Xuất thống kê giảng viên dạng CSV.

    - **sort_by**: Tiêu chí sắp xếp (doanh thu/số học viên/số khóa học)
    - Download file CSV trực tiếp
    """
    import csv
    import io

    await authorization.require_role(["ADMIN"])

    data = await service.get_instructor_export_async(sort_by, from_date, to_date)

    # Tên file
    timestamp = date.today().strftime("%Y%m%d")
    filename = f"instructors_{sort_by}_{timestamp}.csv"

    # StreamingResponse với BOM cho Excel
    stream = io.BytesIO()
    stream.write(b"\xef\xbb\xbf")  # BOM

    text_stream = io.StringIO()
    if data:
        # Header tiếng Việt
        header_map = {
            "stt": "STT",
            "instructor_id": "ID Giảng viên",
            "name": "Họ tên",
            "email": "Email",
            "courses": "Số khóa học",
            "students": "Số học viên",
            "rating": "Đánh giá",
            "total_revenue": "Tổng doanh thu (VND)",
        }

        original_keys = list(data[0].keys())
        writer = csv.DictWriter(text_stream, fieldnames=original_keys)
        # Write header tiếng Việt
        writer.writerow(
            dict(zip(original_keys, [header_map.get(k, k) for k in original_keys]))
        )
        writer.writerows(data)

    stream.write(text_stream.getvalue().encode("utf-8"))
    stream.seek(0)

    from fastapi.responses import StreamingResponse

    return StreamingResponse(
        stream,
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ================================================================
# 20. INSTRUCTOR DETAIL (Dynamic route - phải đặt cuối cùng)
# ================================================================
@router.get("/instructors/{instructor_id}")
async def get_instructor_detail(
    instructor_id: str,
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
):
    """
    Chi tiết thống kê của một giảng viên.

    - Tổng thu nhập
    - Số học viên
    - Rating
    - Top khóa học
    """
    from uuid import UUID

    await authorization.require_role(["ADMIN"])
    return await service.get_instructor_detail_async(UUID(instructor_id))


# ================================================================
# 21. COMPREHENSIVE REPORT EXPORT (Xuất báo cáo toàn diện - Excel)
# ================================================================
@router.get("/export/comprehensive")
async def get_comprehensive_report(
    authorization: AuthorizationService = Depends(AuthorizationService),
    service: StatisticsService = Depends(StatisticsService),
    from_date: Optional[date] = Query(None, description="Ngày bắt đầu (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="Ngày kết thúc (YYYY-MM-DD)"),
    format: str = Query("xlsx", description="Định dạng xuất: xlsx|json"),
):
    """
    Xuất báo cáo toàn diện dạng Excel hoặc JSON.

    Bao gồm:
    - Tổng quan (Overview)
    - Doanh thu theo tháng
    - Doanh thu theo danh mục
    - Thống kê người dùng
    - Thống kê khóa học
    - Top khóa học
    - Thống kê giảng viên
    - Top giảng viên
    - Thống kê tài chính
    - Thống kê hoạt động
    """
    import io

    await authorization.require_role(["ADMIN"])

    report = await service.get_comprehensive_report_async(from_date, to_date)

    # Nếu format là JSON, trả về JSON
    if format == "json":
        return report

    # Xuất Excel với openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to JSON if openpyxl not installed
        return report

    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def create_summary_sheet(ws, title, data: dict):
        """Tạo sheet dạng key-value cho summary"""
        ws.title = title
        ws.append(["Chỉ số", "Giá trị"])
        for key, value in data.items():
            ws.append([key, value])

        # Style header
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Style data
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

        # Auto width
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30

    def create_table_sheet(ws, title, data: list):
        """Tạo sheet dạng bảng cho list of dicts"""
        ws.title = title
        if not data:
            ws.append(["Không có dữ liệu"])
            return

        # Header
        headers = list(data[0].keys())
        ws.append(headers)

        # Data
        for row in data:
            ws.append([row.get(h, "") for h in headers])

        # Style header
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Style data
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

        # Auto width
        for col in range(1, ws.max_column + 1):
            max_length = max(
                len(str(ws.cell(row=row, column=col).value or ""))
                for row in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

    # Sheet 1: Tổng quan
    ws_overview = wb.active
    create_summary_sheet(ws_overview, "Tổng quan", report["overview"])

    # Sheet 2: Doanh thu tổng
    ws_rev_summary = wb.create_sheet()
    create_summary_sheet(ws_rev_summary, "Doanh thu", report["revenue_summary"])

    # Sheet 3: Doanh thu theo tháng
    ws_rev_monthly = wb.create_sheet()
    create_table_sheet(ws_rev_monthly, "DT theo tháng", report["revenue_monthly"])

    # Sheet 4: Doanh thu theo danh mục
    ws_rev_cat = wb.create_sheet()
    create_table_sheet(ws_rev_cat, "DT theo danh mục", report["revenue_by_category"])

    # Sheet 5: Người dùng
    ws_user = wb.create_sheet()
    create_summary_sheet(ws_user, "Người dùng", report["user_summary"])

    # Sheet 6: Người dùng theo vai trò
    ws_user_role = wb.create_sheet()
    create_table_sheet(ws_user_role, "ND theo vai trò", report["users_by_role"])

    # Sheet 7: Khóa học
    ws_course = wb.create_sheet()
    create_summary_sheet(ws_course, "Khóa học", report["course_summary"])

    # Sheet 8: Khóa học theo cấp độ
    ws_course_level = wb.create_sheet()
    create_table_sheet(ws_course_level, "KH theo cấp độ", report["courses_by_level"])

    # Sheet 9: Top khóa học
    ws_top_course = wb.create_sheet()
    create_table_sheet(ws_top_course, "Top khóa học", report["top_courses"])

    # Sheet 10: Giảng viên
    ws_instructor = wb.create_sheet()
    create_summary_sheet(ws_instructor, "Giảng viên", report["instructor_summary"])

    # Sheet 11: Top giảng viên
    ws_top_instructor = wb.create_sheet()
    create_table_sheet(ws_top_instructor, "Top giảng viên", report["top_instructors"])

    # Sheet 12: Tài chính
    ws_finance = wb.create_sheet()
    create_summary_sheet(ws_finance, "Tài chính", report["finance_summary"])

    # Sheet 13: Hoạt động
    ws_activity = wb.create_sheet()
    create_summary_sheet(ws_activity, "Hoạt động", report["activity_summary"])

    # Save to BytesIO
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Generate filename
    timestamp = date.today().strftime("%Y%m%d")
    filename = f"bao_cao_toan_dien_{timestamp}.xlsx"

    from fastapi.responses import StreamingResponse

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
